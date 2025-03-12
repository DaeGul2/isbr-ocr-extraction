import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import zipfile

from functions.parse_text import parse_ocr_text
from functions.classify_document import classify_document
from functions.test_chobon_extract import extract_info_from_chobon
from functions.test_dungbon_extract import extract_info_from_dungbon
from functions.test_geonbojakyeock_extract import extract_info_from_geonbojakyeock
from functions.test_grade_extract import extract_info_from_grade
from functions.test_graduation_extract import extract_info_from_graduation
from functions.test_nps_extract import extract_info_from_nps
from functions.test_toeic_extract import extract_info_from_toeic
from functions.test_toss_extract import extract_info_from_toss

# 🔹 [1] Input 데이터 로드
input_path = "./마사회_2025_신입_ocr_results.xlsx"
df_input = pd.read_excel(input_path)

# 🔹 [2] Output Excel 파일 설정
output_path = "./마사회_2025_신입_output.xlsx"

# 🔹 [3] 워크북 로드 (있으면 열기, 없으면 새로 생성)
if os.path.exists(output_path):
    try:
        book = load_workbook(output_path)
    except zipfile.BadZipFile:
        print("손상된 Excel 파일입니다. 새 파일을 생성합니다.")
        book = Workbook()
else:
    print("Excel 파일이 없습니다. 새 파일을 생성합니다.")
    book = Workbook()

# 🔹 [4] 시트별 데이터 저장을 위한 딕셔너리
sheets_data = {}

# 🔹 [5] 각 행을 순회하며 OCR 데이터 처리
for index, row in df_input.iterrows():
    exam_number, name, ocr_text = row["수험번호"], row["이름"], row["ocr_text"]
    parsed_text = parse_ocr_text(ocr_text)
    
    for text in parsed_text:
        doc_type = classify_document(text)
        
        if doc_type:
            data_extractor = {
                '등본': extract_info_from_dungbon,
                '초본': extract_info_from_chobon,
                '건강보험자격득실': extract_info_from_geonbojakyeock,
                '국민연금가입자증명': extract_info_from_nps,
                '토익': extract_info_from_toeic,
                '토스': extract_info_from_toss,
                '성적증명서': extract_info_from_grade,
                '졸업증명서': extract_info_from_graduation
            }.get(doc_type)

            if data_extractor:
                extracted_data = data_extractor(name, text)
                
                # 🔹 시트 이름 = 문서 유형 (doc_type)
                sheet_name = doc_type

                # 🔹 시트별 데이터 저장을 위한 리스트 생성 (없으면 새로 만듦)
                if sheet_name not in sheets_data:
                    sheets_data[sheet_name] = []

                # 🔹 행 데이터 구성 (검출_원본 제외)
                filtered_data = {k: v for k, v in extracted_data.items() if k != "검출_원본"}
                row_data = [exam_number, name] + [', '.join(values) for values in filtered_data.values()]

                # 🔹 해당 시트의 데이터 리스트에 추가
                sheets_data[sheet_name].append(row_data)

# 🔹 [6] 중복 제거 없이 결과 저장
for sheet_name, data in sheets_data.items():
    # 🔹 가장 긴 행의 길이를 기준으로 컬럼 개수 조정
    max_cols = max(len(row) for row in data)

    # 🔹 컬럼 리스트 동적으로 생성
    default_columns = ["수험번호", "이름"]
    extra_columns = [f"항목_{i+1}" for i in range(max_cols - len(default_columns))]
    column_names = default_columns + extra_columns

    # 🔹 데이터프레임 변환
    df = pd.DataFrame(data, columns=column_names)

    # 🔹 시트가 없으면 생성
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
    else:
        sheet = book.create_sheet(title=sheet_name)
        sheet.append(df.columns.tolist())  # 🔹 헤더 추가

    # 🔹 데이터 저장
    for row in df.itertuples(index=False, name=None):
        sheet.append(row)

# 🔹 [7] 저장하고 파일 닫기
book.save(output_path)
book.close()
print("✅ OCR 분석 결과 저장 완료! (중복 제거 없음)")
